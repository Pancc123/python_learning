# -*- coding=utf-8 -*
# @Time :  2021-1-22 15:07
# @Author : ChengCheng Pan


import pandas as pd
import xlwt,csv
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook,Workbook,styles
from openpyxl.styles import Color, Font, Alignment
from openpyxl.utils import get_column_letter


def csv_to_xlsx_pd(old_file,new_file):
    csv = pd.read_csv(old_file, encoding='utf-8')
    print(csv)
    csv.to_excel(new_file, sheet_name='data')


# 设置excel 居中，自动换行
def modify_excel_style(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    align = Alignment(vertical='center',wrapText=True)
    nrows=ws.max_row
    ncols=ws.max_column
    print(ncols,nrows)
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=i + 1, column=j + 1).alignment = align

    # def adjust_column_dimension(ws, min_row=1, min_col=1, max_col=ncols):
    #     column_widths = []
    #     for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
    #         for cell in col:
    #             value = cell.value
    #             if value is not None:
    #                 if isinstance(value, str) is False:
    #                     value = str(value)
    #                 try:
    #                     column_widths[i] = max(column_widths[i], len(value))
    #                 except IndexError:
    #                     column_widths.append(len(value))
    #     for i, width in enumerate(column_widths):
    #         col_name = get_column_letter(min_col + i)
    #         value = column_widths[i] + 2
    #         ws.column_dimensions[col_name].width = value
        # 获取每一列的内容的最大宽度
    # i = 0
    # # 每列
    # col_width=[]
    # for col in ws.columns:
    #     # 每行
    #     for j in range(ncols+1):
    #         if j == 0:
    #             # 数组增加一个元素
    #             col_width.append(len(str(col[j].value)))
    #         else:
    #             print(col_width)
    #             print(i)
    #             # 获得每列中的内容的最大宽度
    #             if col_width[0] < len(str(col[j].value)):
    #                 col_width[0] = len(str(col[j].value))
    #         i = i + 1
    wb.save(path)


# 手动修改行宽
def modify_excel_width(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20
    wb.save(path)



def modify_excel_style1(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    nrows = ws.max_row
    ncols = ws.max_column
    col_width = []
    for i in range(1, nrows + 1):
        for j in range(1, ncols + 1):
            # # 方法一：
            # if j == 1:
            #     # 数组增加一个元素
            #     # col_width.append(len(str(ws.cell(i,j).value)))
            #     col_width=len(str(ws.cell(i,j).value))
            # else:
            #     # print(col_width)
            #     # print(i)
            #     # 获得每列中的内容的最大宽度
            #     if col_width < len(str(ws.cell(i,j).value)):
            #         col_width = len(str(ws.cell(i,j).value))
            # 方法二：

            col_width.append(len(str(ws.cell(i, j).value)))
        # 设置列宽
        # 根据列的数字返回字母
        # print(col_width)
        # print(max(col_width))
        col_letter = get_column_letter(i)
        # 当宽度大于100，宽度设置为100
        if max(col_width) > 50:
            ws.column_dimensions[col_letter].width = 50
        # 只有当宽度大于10，才设置列宽
        elif max(col_width) > 10:
            ws.column_dimensions[col_letter].width = max(col_width) + 2
    wb.save(path)


def csv_to_xlsx(csvfile,outfile):
    with open(csvfile, encoding='utf-8') as fc:
        r_csv = csv.reader(fc)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('sheet1')  # 创建一个sheet表格
        i = 0
        j = 0
        for line in r_csv:
            for v in line:
                sheet.write(i,j,v)
                j = j + 1
            i = i + 1
        workbook.save(outfile)  # 保存Excel

if __name__ == '__main__':
    old_file=r'D:\N3-K网管-Kddi新增功能-用例.csv'
    new_file=r'D:\N3-K网管-Kddi新增功能-用例1.xlsx'
    # old_file=r'D:\N3-K网管-所有用例.csv'
    # new_file=r'D:\N3-K网管-所有用例1.xlsx'
    csv_to_xlsx_pd(old_file,new_file)
    sheet_name = "data"
    modify_excel_style(new_file)
    #modify_excel_style1(path)
    modify_excel_width(new_file)