# -*- coding=utf-8 -*
# @Time :  2021-1-22 15:07
# @Author : ChengCheng Pan
import pandas as pd
import xlwt,csv
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook,Workbook,styles
from openpyxl.styles import Border, Color, Font, Alignment,Side
from openpyxl.utils import get_column_letter
from io import StringIO
c_path = r'D:\N3-K网管-功能需求迭代二-用例.csv' # 原csv文件
x_path = r'D:\N3-K网管-功能需求迭代二-用例1.xlsx'  # 路径中的xls文件在调用to_excel时会自动创建


def csv_to_xlsx_pd(old_file,new_file):
    csv = pd.read_csv(old_file, encoding='utf-8')
    print(csv)
    csv.to_excel(new_file, sheet_name='data')


# 禅道用例csv转xlsx
def csv_to_xlsx(c_path, x_path):
    with open(c_path, 'r', encoding='utf-8') as f:
        data = f.read()
        # print(data)
        data_file = StringIO(data)
        # print(data_file)
        csv_reader = csv.reader(data_file)
        # print(csv_reader)
        list_csv = []

        for row in csv_reader:
            list_csv.append(row)
            df_csv = pd.DataFrame(list_csv).applymap(str)
            # print(df_csv)

    writer = pd.ExcelWriter(x_path)
    # 写入Excel
    df_csv.to_excel(excel_writer=writer,index=False,header=False,sheet_name='N3-K网管-功能需求迭代二-用例' )
    writer.save()


def csv_to_xlsx1(csvfile,outfile):
    with open(csvfile, encoding='utf-8') as fc:
        r_csv = csv.reader(fc)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('sheet1')  # 创建一个sheet表格
        i = 0
        j = 0
        for line in r_csv:
            for v in line:
                sheet.write(i,j,v)
                j = j + 1
            i = i + 1
        workbook.save(outfile)  # 保存Excel


# 设置excel 居中，自动换行
def set_middle(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    align = Alignment(vertical='center',wrapText=True)  # 列居中，自动换行
    align1 = Alignment(horizontal="center",vertical='center', wrapText=True)  # 列居中，自动换行
    nrows = ws.max_row
    ncols = ws.max_column
    print(ncols,nrows)
    for i in range(nrows):
        for j in range(ncols):
            # 设置第一行行列居中
            if i == 0:
                ws.cell(row=i + 1, column=j + 1).alignment = align1
            else:
                ws.cell(row=i + 1, column=j + 1).alignment = align

    # def adjust_column_dimension(ws, min_row=1, min_col=1, max_col=ncols):
    #     column_widths = []
    #     for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
    #         for cell in col:
    #             value = cell.value
    #             if value is not None:
    #                 if isinstance(value, str) is False:
    #                     value = str(value)
    #                 try:
    #                     column_widths[i] = max(column_widths[i], len(value))
    #                 except IndexError:
    #                     column_widths.append(len(value))
    #     for i, width in enumerate(column_widths):
    #         col_name = get_column_letter(min_col + i)
    #         value = column_widths[i] + 2
    #         ws.column_dimensions[col_name].width = value
        # 获取每一列的内容的最大宽度
    # i = 0
    # # 每列
    # col_width=[]
    # for col in ws.columns:
    #     # 每行
    #     for j in range(ncols+1):
    #         if j == 0:
    #             # 数组增加一个元素
    #             col_width.append(len(str(col[j].value)))
    #         else:
    #             print(col_width)
    #             print(i)
    #             # 获得每列中的内容的最大宽度
    #             if col_width[0] < len(str(col[j].value)):
    #                 col_width[0] = len(str(col[j].value))
    #         i = i + 1
    wb.save(path)


# 手动修改行宽
def modify_excel_width(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20
    wb.save(path)


def set_cols_width(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    nrows = ws.max_row
    ncols = ws.max_column
    # col_width = []
    # for i in range(1, nrows + 1):
    #     col_width = []
    #     for j in range(1, ncols + 1):
    #         # # 方法一：
    #         # if j == 1:
    #         #     # 数组增加一个元素
    #         #     # col_width.append(len(str(ws.cell(i,j).value)))
    #         #     col_width=len(str(ws.cell(i,j).value))
    #         # else:
    #         #     # print(col_width)
    #         #     # print(i)
    #         #     # 获得每列中的内容的最大宽度
    #         #     if col_width < len(str(ws.cell(i,j).value)):
    #         #         col_width = len(str(ws.cell(i,j).value))
    #         # 方法二：
    #         length = len(str(ws.cell(i, j).value))
    #         col_width.append(length)
    for j in range(1,ncols+1):
        col_width = []
        for i in range(1,nrows+1):
            length = len(str(ws.cell(i, j).value))
            col_width.append(length)

        # 设置列宽
        # print(col_width)
        # print(max(col_width))
        col_letter = get_column_letter(j)  # 根据列的数字返回字母
        # ws.column_dimensions[col_letter].width =max(col_width)+2
        # 当宽度大于100，宽度设置为100
        if max(col_width) > 50:
            ws.column_dimensions[col_letter].width = 50
        # 只有当宽度大于10，才设置列宽
        elif max(col_width) > 10:
            ws.column_dimensions[col_letter].width = max(col_width) + 2
        else:
            ws.column_dimensions[col_letter].width = max(col_width) + 6
    # 删除列，从第几列开始几列
    # ws.delete_cols(2, 1)
    # ws.delete_cols(3, 1)
    wb.save(path)


# 设置边框线
def set_border(path):
   wb = load_workbook(path)
   ws = wb[wb.sheetnames[0]]
   line_style = Side(border_style='thin')  # 边框线条选择，实线
   line_style1 = Side(border_style='medium')  # 边框线条选择，实线
   border = Border(top=line_style, right=line_style, left=line_style, bottom=line_style)  # 单元格边框格式
   border1 = Border(top=line_style1, right=line_style1, left=line_style1, bottom=line_style1)
   nrows = ws.max_row
   ncols = ws.max_column
   for i in range(nrows):
       for j in range(ncols):
           # 设置第一行行列居中
           if i == 0:
               ws.cell(row=i + 1, column=j + 1).border = border1
           else:
               ws.cell(row=i + 1, column=j + 1).border = border
   wb.save(path)


class SetExcel():
    def __init__(self, c_path, x_path):
        self.c_path = c_path  # 源文件路径
        self.x_path = x_path  # 新文件路径
        wb = load_workbook(c_path)
        self.ws = wb[wb.sheetnames[0]]
        self.nrows = self.ws.max_row
        self.ncols = self.ws.max_column

    def set_align(self,cell_obj): # cell_obj单元格对象
        align = Alignment(horizontal="center", vertical='center', wrapText=True)
        cell_obj.alignment = align

    def set_cols_width(self,col_id,width): # col_id:列号 width：宽度
        col_letter = get_column_letter(col_id)
        self.ws.column_dimensions[col_letter].width=width

    def delete_cols(self,col_id,num):
        self.ws.delete_cols(col_id, num)

    def set_border(self,border_style,cell_obj):
        line_style = Side(border_style=border_style)
        border = Border(top=line_style, right=line_style, left=line_style, bottom=line_style)
        cell_obj.border=border


if __name__ == '__main__':

    # old_file=r'D:\N3-K网管-功能需求迭代二-用例.csv'
    # new_file=r'D:\N3-K网管-功能需求迭代二-用例.xlsx'
    # # old_file=r'D:\N3-K网管-所有用例.csv'
    # # new_file=r'D:\N3-K网管-所有用例1.xlsx'
    # csv_to_xlsx_pd(old_file,new_file)
    # sheet_name = "data"
    # set_middle(new_file)
    # modify_excel_width(new_file)
    # 禅道用例csv处理
    csv_to_xlsx(c_path, x_path)
    set_cols_width(x_path)
    set_middle(x_path)
    set_border(x_path)
