# -*- coding=utf-8 -*
# @Time :  2021-1-26 19:27
# @Author : ChengCheng Pan
import pandas as pd
import xlwt,csv
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook,Workbook,styles
from openpyxl.styles import Color, Font, Alignment
from openpyxl.utils import get_column_letter

new_file = r'D:\N3-K网管-Kddi新增功能-用例1.xlsx'
# 打开已有的1个excel，之后保存
def open_excel(path):
    def deco_func(fun):
        def wrapper(*args, **kwargs):
            wb = load_workbook(path)
            ws = wb[wb.sheetnames[0]]
            wb.save(path)
            f = fun(*args, **kwargs)
            # 通过索引得到sheet
    # sheetnames = wb.sheetnames
    # print(sheetnames)
    # print(wb["data"]) 以sheetname做索引查询sheet
    # ws = wb[wb.sheetnames[0]]


        return wrapper
    return deco_func


@open_excel(new_file)
def modify_excel_width(path):
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.column_dimensions['C'].width = 12 # 设置行宽
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20
    wb.save(path)


if __name__== "__main__":
    new_file = r'D:\N3-K网管-Kddi新增功能-用例1.xlsx'
    open_excel(new_file)