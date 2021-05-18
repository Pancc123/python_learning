# -*- coding=utf-8 -*
# @Time :  2021-4-21 9:59
# @Author : ChengCheng Pan
from openpyxl import load_workbook,Workbook,styles
import json,os

path1 = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'test_data', 'data1.xlsx').replace('\\', '/')


def write_data_to_xlsx(sheet_name):
    wb = Workbook()
    ws = wb.create_sheet(sheet_name, index=0)
    list1 = ["case_name", "params", "code", "message", "data", "type"]
    ws.append(list1)
    postdata = {"action": "serviceProviderLogin",
                "username": 'admin_pcc',
                "password": '123456',
                "exp": "1"}
    postdata = json.dumps(postdata).encode('utf-8')
    ws.cell(2,2).value=postdata
    font = styles.Font(name=u'宋体', bold=True)
    align = styles.Alignment(horizontal='center', vertical='center')
    #for i in range(2,len(data)+1):
        #for j in range(1,len(data[i-1])+1):
           # ws.cell(row=i,column=j,value=str(data[i-1][j-1]))
    #font = styles.Font(name=u'宋体', bold=True)
    #align = styles.Alignment(horizontal='center', vertical='center')

    #ws.cellstyle('A1', font, align)
    wb.save(path1)


def write_sqldata(data,save_dir):
    wb = Workbook()
    wc = wb.create_sheet("Sqldata",index=0)
    wc.append(['service_name', 'url', 'parameters'], )
    #data1=data[0]
    #fields=data[1]
    #for field in range(1,len(fields)):
    #   wc.cell(row=1,column=field,value=str(fields[field-1]))
    for i in range(2,len(data)+1):
        for j in range(1,len(data[i-1])+1):
            wc.cell(row=i,column=j,value=str(data[i-1][j-1]))
    # wc.append(['service_name','url','parameters'],)
    wb.save(save_dir)


def modify_cell_format(save_dir):
    wt=load_workbook(save_dir)
    sheet=wt.get_sheet_by_name('Sqldata')
    sheet.column_dimensions['A'].width =40
    sheet.column_dimensions['B'].width =25
    sheet.column_dimensions['C'].width =150
    size_font = styles.Font(size=12,bold=True)
    colums=sheet.max_column
    for i in range(1,colums+1):
        sheet.cell(row=1,column=i).font = size_font
    # sheet.column_dimensions[1].font=size_font
    wt.save(save_dir)
