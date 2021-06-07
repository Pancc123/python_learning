# -*- coding=utf-8 -*
# @Time :  2021-5-21 14:23
# @Author : ChengCheng Pan
from openpyxl import load_workbook,Workbook,styles
from openpyxl.styles import Border, colors, Font, Alignment,Side,PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os

# path1 = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'test_data', 'data1.xlsx').replace('\\', '/')
# 生成文件目录


# 创建excel,写入数据
def create_excel(path, sheet_name,data):
    wb = Workbook()
    ws = wb.create_sheet(sheet_name,index=0)
    rows = len(data)
    columns = len(data[0])
    for i in range(1,rows+1):
        for j in range(1,columns+1):
            ws.cell(row=i, column=j, value=data[i-1][j-1])
    wb.save(path)


# 原有文件创建新的sheet
def add_sheet(path,sheet_name,data):
    wb = load_workbook(path)
    ws = wb.create_sheet(sheet_name,index=0)
    rows = len(data)
    columns = len(data[0])
    for i in range(1,rows+1):
        for j in range(1,columns+1):
            ws.cell(row=i, column=j, value=data[i-1][j-1])
    wb.save(path)


# 写数据到原有文件sheet页中
def modify_sheet_data(path,sheet_name,data):  # 还可通过索引wb.sheetnames[0]获取sheet_name
    wb = load_workbook(path)
    ws = wb[sheet_name]
    rows = len(data)
    columns = len(data[0])
    for i in range(1,rows+1):
        for j in range(1,columns+1):
            ws.cell(row=i, column=j, value=data[i-1][j-1])
    wb.save(path)


def read_sheet_data(path,sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    rows = ws.max_row
    columns = ws.max_column
    data_list = list()
    for i in range(1,rows+1):
        row_data = list()
        for j in range(1,columns+1):
            cell_data = ws.cell(i,j).value
            row_data.append(cell_data)
        data_list.append(row_data)
    wb.save(path)
    return data_list


def read_sheet_data1(path,sheet_name):   # 获取行号也可这样 rows_num =len(list(ws.rows))
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for row in ws.rows:     # 也可以是 for column in ws.columns
        row_data = list()
        for cell in row:
            print(cell.value)
            row_data.append(cell.value)
        data_list.append(row_data)
    wb.save(path)
    return data_list


# 设置excel 居中，自动换行
def set_middle(path,sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name] # sheet_name=wb.sheetnames[0]
    align = Alignment(vertical='center',wrapText=True)  # 列居中，自动换行
    align1 = Alignment(horizontal="center",vertical='center', wrapText=True)  # 列居中，自动换行
    nrows = ws.max_row
    ncols = ws.max_column
    print(ncols,nrows)
    for i in range(nrows):
        for j in range(ncols):
            # 设置第一行行列居中
            if i == 0:
                ws.cell(row=i + 1, column=j + 1).alignment = align1
            else:
                ws.cell(row=i + 1, column=j + 1).alignment = align

    # def adjust_column_dimension(ws, min_row=1, min_col=1, max_col=ncols):
    #     column_widths = []
    #     for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
    #         for cell in col:
    #             value = cell.value
    #             if value is not None:
    #                 if isinstance(value, str) is False:
    #                     value = str(value)
    #                 try:
    #                     column_widths[i] = max(column_widths[i], len(value))
    #                 except IndexError:
    #                     column_widths.append(len(value))
    #     for i, width in enumerate(column_widths):
    #         col_name = get_column_letter(min_col + i)
    #         value = column_widths[i] + 2
    #         ws.column_dimensions[col_name].width = value
        # 获取每一列的内容的最大宽度
    # i = 0
    # # 每列
    # col_width=[]
    # for col in ws.columns:
    #     # 每行
    #     for j in range(ncols+1):
    #         if j == 0:
    #             # 数组增加一个元素
    #             col_width.append(len(str(col[j].value)))
    #         else:
    #             print(col_width)
    #             print(i)
    #             # 获得每列中的内容的最大宽度
    #             if col_width[0] < len(str(col[j].value)):
    #                 col_width[0] = len(str(col[j].value))
    #         i = i + 1
    wb.save(path)


# 手动修改列宽
def modify_excel_width(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20
    wb.save(path)


# 自动设置列宽
def set_cols_width(path):
    wb = load_workbook(path)
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    nrows = ws.max_row
    ncols = ws.max_column
    # col_width = []
    # for i in range(1, nrows + 1):
    #     col_width = []
    #     for j in range(1, ncols + 1):
    #         # # 方法一：
    #         # if j == 1:
    #         #     # 数组增加一个元素
    #         #     # col_width.append(len(str(ws.cell(i,j).value)))
    #         #     col_width=len(str(ws.cell(i,j).value))
    #         # else:
    #         #     # print(col_width)
    #         #     # print(i)
    #         #     # 获得每列中的内容的最大宽度
    #         #     if col_width < len(str(ws.cell(i,j).value)):
    #         #         col_width = len(str(ws.cell(i,j).value))
    #         # 方法二：
    #         length = len(str(ws.cell(i, j).value))
    #         col_width.append(length)
    for j in range(1,ncols+1):
        col_width = []
        for i in range(1,nrows+1):
            length = len(str(ws.cell(i, j).value))
            col_width.append(length)

        # 设置列宽
        # print(col_width)
        # print(max(col_width))
        col_letter = get_column_letter(j)  # 根据列的数字返回字母
        # ws.column_dimensions[col_letter].width =max(col_width)+2
        # 当宽度大于100，宽度设置为100
        if max(col_width) > 50:
            ws.column_dimensions[col_letter].width = 50
        # 只有当宽度大于10，才设置列宽
        elif max(col_width) > 10:
            ws.column_dimensions[col_letter].width = max(col_width) + 2
        else:
            ws.column_dimensions[col_letter].width = max(col_width) + 6
    # 删除列，从第几列开始几列
    # ws.delete_cols(2, 1)
    # ws.delete_cols(3, 1)
    wb.save(path)


# 设置边框线
def set_border(path):
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    line_style = Side(border_style='thin')  # 边框线条选择，实线
    line_style1 = Side(border_style='medium')  # 边框线条选择，实线
    border = Border(top=line_style, right=line_style, left=line_style, bottom=line_style)  # 单元格边框格式
    border1 = Border(top=line_style1, right=line_style1, left=line_style1, bottom=line_style1)
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(1,nrows+1):
        for j in range(1,ncols+1):
            #设置第一行行列居中
            if i == 0:
                ws.cell(row=i, column=j).border = border1
            else:
                ws.cell(row=i, column=j).border = border
    wb.save(path)


# 设置填充色
def set_fill(path,sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]  # sheet_name=wb.sheetnames[0]
    # 方法一
    # fille = PatternFill(patternType="solid",fgColor="FFBB02",) # solid 是填充实色的意思;fgColor好像还可以这样表示：fgColor=colors.RED
    # 方法二
    fille = PatternFill(patternType="solid", fgColor=colors.BLUE, )
    nrows = ws.max_row
    ncols = ws.max_column
    print(nrows,ncols)
    for i in range(1,nrows+1):
        for j in range(1,ncols+1):
            if i == 0:
                pass
            else:
                ws.cell(i,j).fill = fille
    wb.save(path)


# 设置批注
def set_notation(path,sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]  # sheet_name=wb.sheetnames[0]
    comment = Comment('hello world', 'yunan')  # 插入批注
    ws.cell(2, 2).comment = comment
    # ws["A1"].comment = comment
    wb.save(path)


# 设置字体
def set_word_style(path,sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]  # sheet_name=wb.sheetnames[0]
    font1 = styles.Font(name=u'宋体',size=12, bold=True)
    ws.cell(row=1, column=1).font = font1
    wb.save(path)


class Excel:
    def __init__(self, c_path, x_path,sheet_name):
        self.c_path = c_path  # 源文件路径
        self.x_path = x_path  # 新文件路径
        self.sheet_name=sheet_name

    def create_sheet(self):
        self.wb = Workbook()
        self.ws = self.wb.create_sheet(sheet_name, index=0)

    def load_excel(self):
        self.wb = load_workbook(self.c_path)
        self.ws = self.wb[sheet_name]
        self.nrows = ws.max_row
        self.ncols = ws.max_column

    def write_to_excel(self,data):
        rows = len(data)
        columns = len(data[0])
        for i in range(1, rows + 1):
            for j in range(1, columns + 1):
                self.ws.cell(row=i, column=j, value=data[i - 1][j - 1])

    def set_align(self,cell_obj): # cell_obj单元格对象
        align = Alignment(horizontal="center", vertical='center', wrapText=True)
        cell_obj.alignment = align

    def set_cols_width(self,col_id,width): # col_id:列号 width：宽度
        col_letter = get_column_letter(col_id)
        self.ws.column_dimensions[col_letter].width=width

    def delete_cols(self,col_id,num):
        self.ws.delete_cols(col_id, num)

    def set_border(self,border_style,cell_obj):
        line_style = Side(border_style=border_style)
        border = Border(top=line_style, right=line_style, left=line_style, bottom=line_style)
        cell_obj.border=border

    def save_excel(self, path1):
        self.wb.save(path1)


if __name__ == '__main__':
    data_list = [[1,2,3],[4,5,6]]
    path = r'D://xlsx_test.xlsx'
    sheet_name = 'test'
    # create_excel(path,sheet_name,data_list)
    # set_fill(path,sheet_name)
    wb = load_workbook(path)
    ws = wb[sheet_name]  # sheet_name=wb.sheetnames[0]
    font1 = styles.Font(name=u'宋体',size=12, bold=True)
    ws.cell(row=1, column=1).font = font1
    wb.save(r'D://xlsx_test1.xlsx')
