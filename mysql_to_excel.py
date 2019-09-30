from  mysql_db import DB,getSQLCONFIG
from mysql_query import query_database
import os
from openpyxl import Workbook,load_workbook,styles

base_dir = str(os.path.dirname(os.path.dirname(__file__)))
base_dir = base_dir+'/'+'SQLconfig.config'
print(base_dir)
config = getSQLCONFIG(r'D:/python_test/SQLconfig.config','Database3')
host = config[1]
username = config[2]
password = config[3]
db_name = config[0]
username = username.replace("'","")
password = password.replace("'","")
db_name = db_name.replace("'","")
host = host.replace("'","")

def query_data(sql):
    db = DB(host, username, password, db_name)
    data = query_database(db, sql)
    return data


def write_sqldata(data,save_dir):
    wb = Workbook()
    wc = wb.create_sheet("Sqldata",index=0)

    wc.append(['service_name', 'url', 'parameters'], )
    #data1=data[0]
    #fields=data[1]
    #for field in range(1,len(fields)):
    #   wc.cell(row=1,column=field,value=str(fields[field-1]))
    for i in range(2,len(data)+1):
        for j in range(1,len(data[i-1])+1):
            wc.cell(row=i,column=j,value=str(data[i-1][j-1]))
    #wc.append(['service_name','url','parameters'],)
    wb.save(save_dir)


def modify_cell_format(save_dir):
    wt=load_workbook(save_dir)
    sheet=wt.get_sheet_by_name('Sqldata')
    sheet.column_dimensions['A'].width =40
    sheet.column_dimensions['B'].width =25
    sheet.column_dimensions['C'].width =150
    size_font=styles.Font(size=12,bold=True)
    colums=sheet.max_column
    for i in range(1,colums+1):
        sheet.cell(row=1,column=i).font=size_font
    #sheet.column_dimensions[1].font=size_font
    wt.save(save_dir)


if __name__ =='__main__':
    sql = "select service_name,url,parameters from openapi_service"
    save_dir="D:/openapi.xlsx"
    t=query_data(sql)
    write_sqldata(t,save_dir)
    modify_cell_format(save_dir)